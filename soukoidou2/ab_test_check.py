import platform
from numpy import empty
import pandas as pd
import os
from typing import Dict, cast
import openpyxl
from openpyxl.cell.cell import Cell

import sys

from fetch_data import IFetchData
from recorder import Recorder
from create_koito_coa import CreateKoitoCoa


class ABTestCheck:

    def __init__(self, fetchKoitoKensa: IFetchData,
                 createKoitoCoa: CreateKoitoCoa,
                 recorder: Recorder)-> None:

        self._createKoitoCoa: CreateKoitoCoa = createKoitoCoa
        self._recorder = recorder

        self._path = r'\\192.168.1.247\共有\技術課ﾌｫﾙﾀﾞ\品質検査\小糸B試験管理ｼｰﾄ.xlsx'
        if platform.system() == 'Linux':
            self._path = r'/mnt/public/技術課ﾌｫﾙﾀﾞ/品質検査/小糸B試験管理ｼｰﾄ.xlsx'

        self._ab_check_df = pd.read_excel(self._path, sheet_name='matome', skiprows=4)
        hinbanTaiou_df = pd.read_excel(self._path, sheet_name='hinbantaiou')
        # hinbans = {'S4-K706CLA-U': 'K706A', ......}
        # hinmeis = {'K706改C':['S4-K706CLKC-U', 'S4-K706CLKC-4-U']....}を作る
        self._hinbans: Dict = dict(zip(hinbanTaiou_df['hinban'], 
                                                    hinbanTaiou_df['hinmei']))
        self._hinmeis: Dict = self.make_hinmeis(hinbanTaiou_df[['hinban', 'hinmei']])
            
        # AB試験で合格した全データのデータフレーム
        '''
        IDOU ={'01':'中', '02':'済', '03':'特'}
        SHIKEN = {'01':'A', '02':'B', '03':'-'}
        '''
        passed_koitos_all: pd.DataFrame = fetchKoitoKensa.fetch_data()
        # 今回の小糸検査(済でない合格品)
        self._passed_koitos_thistime = cast(pd.DataFrame, 
                      passed_koitos_all[passed_koitos_all['IDOU']!='02'].copy())

        if not self._passed_koitos_thistime.empty:
        # 済の小糸検査合格品
            self._passed_koitos_sumi = cast(pd.DataFrame, 
                      passed_koitos_all[passed_koitos_all['IDOU']=='02'].copy())
        # 前回lotを探して入力
            self._passed_koitos_thistime['lastLot'] = \
            self._passed_koitos_thistime.apply(self.find_koito_lastLot, axis=1)
        # 前回lotのcountを入力
            self._passed_koitos_thistime['count_lastTime'] = \
       self._passed_koitos_thistime.apply(self.find_koito_lastLot_count, axis=1)
        

        '''
        1. 今回合格した小糸むけ製品 _passed_koitos_thistime
        2. これまでの小糸向け製品 passed_koitos_all
        3. 前回ロットを探す passed_koito_last
        4. _ab_check_dfで前回ロットが何番目かを求める
        5. 今回合格品のAB判定
        '''

    def is_empty_passed_koitos_thistime(self)-> bool:
        '''空だったらTrue'''
        if self._passed_koitos_thistime.empty:
            txt = '<小糸AB試験チェック結果> \n' \
                  '~ 小糸AB試験はありません。~'
            self._recorder.out_log(txt, '\n')
            self._recorder.out_file(txt, '\n')
        return self._passed_koitos_thistime.empty
        

    def check_is_abTest_ok(self)-> bool:
        '''
        itertuples()を使うと属性アクセスでpyrighが警告を出す
        ので、iterrows()を使った
        '''
        is_abTest_ok = True

        txt: str = '\n AB試験チェックオッケーです'
        As = [1, 2, 3, 5]
        Bs = [4]
        for _, row in self._passed_koitos_thistime.iterrows():

            if row['count_lastTime'] == 0:
                return False
            if row['SHIKEN'] == '01' and row['count_lastTime'] not in As:
                is_abTest_ok =  False
            if row['SHIKEN'] == '02' and row['count_lastTime'] not in Bs:
                is_abTest_ok =  False

        self._recorder.out_log_df(self._passed_koitos_thistime,
                                 '<小糸AB試験チェック結果>')
        self._recorder.out_file_from_df(self._passed_koitos_thistime, 
                                  '<小糸AB試験チェック結果>')

        if not is_abTest_ok:
            txt = '\n xxxxxx AB試験チェックNGです xxxxxx\n' \
                  'プログラムを中断します。'
            self._recorder.out_log(txt, '\n')
            self._recorder.out_file(txt, '\n')
            sys.exit(1)

        self._recorder.out_log(txt, '\n')
        self._recorder.out_file(txt, '\n')
        
        return is_abTest_ok


    def find_koito_lastLot_count(self, row)-> int:
        count:int = 0 # countは1~5
        if row['lastLot'] == 'notFound':
            return count

        lot = row['LOT']
        hinban =  row['Hinban']
        hinmei = self._hinbans[hinban] # K706改B etc
        lastLot = row['lastLot'][:10] # 詰め替え識別Noは除く
        # matching_indices = [0, 3, 10] など、lastLotと一致したindexのリストが返る 
        matching_indices = self._ab_check_df.index[self._ab_check_df[hinmei]
                                    .str.contains(lastLot, na=False)].tolist()
        if not matching_indices:
            return count

        index: int = 0
        index = matching_indices[0]
        count = self._ab_check_df.loc[index, '1～5'] 

        return count


    def find_koito_lastLot(self, row)-> str:
        lastLot: str = 'notFound'
        hinban =  row['Hinban']
        try:
            hinmei = self._hinbans[hinban] # K706改B etc
        except:
            return lastLot # notFound

        hinbans = self._hinmeis[hinmei] # ['S4-K706KB-U', 'S4-K706KB-4-U'] etc
        # 前回lotを求める
        df = cast(pd.DataFrame, 
         self._passed_koitos_sumi[self._passed_koitos_sumi['Hinban'].isin(hinbans)])
        if not df.empty:
            lastLot = cast(str, cast(pd.Series, df['LOT']).iloc[-1])
            # 最終行のlotを取得

        return lastLot


    def make_hinmeis(self, df)-> Dict:
        # hinmeiが_hinmeisに無かったら新たにlist作って、あったらappendする 
        # dfそのまま回してもよいが、列順を固定するためにdf[['hinban', 'hinmei']]にする
        hinmeis: Dict = {}
        for row in df.itertuples(index=False):
            if row[1] in hinmeis:
                hinmeis[row[1]].append(row[0])
                continue

            hinmeis[row[1]] = [row[0]] 

        return hinmeis


    def input_to_BsikenKanriSheet(self)-> None:

        '''この時点でABチェックokは保証されている'''
        # 2. 書き込み権限があるか（使用中でないか）チェック
        try:
            # 既存のファイル名でリネームを試みることで使用中か判定
            os.rename(self._path, self._path)
        except OSError:
            txt = '小糸B試験管理シートが使用中のためプログラムを中断します。'
            self._recorder.out_log(txt, '\n')
            self._recorder.out_file(txt, '\n')
            sys.exit(1)

        wb = openpyxl.load_workbook(self._path)
        ws = wb['matome']
        
        hinmeis_row = 5 # 品名が並んでいる行
        for _, row in self._passed_koitos_thistime.iterrows():
            hinban = row['Hinban']
            hinmei = self._hinbans[hinban]
            lastLot = row['lastLot'][:9]
            lot = row['LOT'][:9]
            input_col = self.get_input_col(hinmei, hinmeis_row, ws)
            input_row = self.get_lastLot_row(cast(str,lastLot), input_col, ws) + 1
            
            if input_col == 0 or input_row == 1:
                txt: str = f'B試験管理シートへの記入({hinmei})失敗です。'
                self._recorder.out_log(txt, '\n')
                self._recorder.out_file(txt, '\n')
                continue  

            # lotを入力
            lot_cell = cast(Cell, ws.cell(input_row, input_col))
            lot_cell.value = cast(str, lot)

            # signを入力
            sign: str = "db"
            sign_cell = cast(Cell, ws.cell(input_row, input_col + 1))
            sign_cell.value = sign
        
        self.save_workbook(wb)

    def save_workbook(self, wb)-> None:
        try:
            # 保存を試みる
            wb.save(self._path)
            txt = '小糸B試験管理シートに記入し、変更を保存しました' 
            self._recorder.out_log(txt, '\n')
            self._recorder.out_file(txt, '\n')
        except PermissionError:
            # ファイルが開かれている場合に発生するエラー
            txt = '小糸B試験管理シートが使用中です。プログラムを中断します。' 
            self._recorder.out_log(txt, '\n')
            self._recorder.out_file(txt, '\n')
            # プログラムを終了する
            sys.exit(1) 
        except Exception as e:
            # その他の予期せぬエラー用
            txt = f'小糸B試験管理シートに変更を保存できません。' \
                  f'プログラムを中断します。{e}' 
            self._recorder.out_log(txt, '\n')
            self._recorder.out_file(txt, '\n')
            sys.exit(1)
    
    def get_input_col(self, hinmei: str, hinmeis_row: int, ws)-> int:
        last_col = 50 # 最終列は多めに50としておく
        input_col:int = 0
        for i in range(1, last_col + 1):
            if ws.cell(hinmeis_row, i).value == hinmei:
                input_col = i
                return input_col
        return input_col

    def get_lastLot_row(self, lastLot: str, input_col: int, ws)-> int:
        last_row = 600 # 最終行
        stt_row = 6
        lastLot_row: int = 0
        for i in range( stt_row, last_row + 1):
            if ws.cell(i, input_col).value == lastLot:
                lastLot_row = i
                return lastLot_row
        return lastLot_row


    def create_koito_coa(self)-> None:
        self._createKoitoCoa.create_koito_coa(self._passed_koitos_thistime)

